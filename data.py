import cv2
import numpy as np
import tkinter as tk
from tkinter import messagebox, ttk
from PIL import Image, ImageTk
from picamera2 import Picamera2
from datetime import datetime
import pandas as pd
import os
import sys
import fcntl
import atexit
from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, Reference
import time

# =============================================================================
# Lock file handling
# =============================================================================
LOCK_FILE = "/tmp/respiratory_therapy.lock"

def check_and_create_lock():
    try:
        # Open or create lock file
        global lock_file
        lock_file = open(LOCK_FILE, 'w')
        
        # Try to acquire an exclusive lock
        fcntl.flock(lock_file, fcntl.LOCK_EX | fcntl.LOCK_NB)
        
        # Register cleanup function
        atexit.register(cleanup_lock)
        
        return True
    except IOError:
        messagebox.showerror("Error", "Another instance of the program is already running.")
        sys.exit(1)

def cleanup_lock():
    try:
        # Release the lock and close/remove the file
        if 'lock_file' in globals():
            fcntl.flock(lock_file, fcntl.LOCK_UN)
            lock_file.close()
        if os.path.exists(LOCK_FILE):
            os.remove(LOCK_FILE)
    except Exception:
        pass

# =============================================================================
# Configuration Constants (for both windows)
# =============================================================================
BACKGROUND_COLOR = "#1a2a5a"  # dark blue
TEXT_COLOR = "white"
FONT_NAME = "Segoe UI"         # Must be defined before use.
RULER_FONT_SIZE = 12

# Main window (ball detection) settings
WINDOW_WIDTH = 700
WINDOW_HEIGHT = 720
HEADER_HEIGHT = 50
CANVAS_WIDTH = WINDOW_WIDTH       # full window width
CANVAS_HEIGHT = WINDOW_HEIGHT - HEADER_HEIGHT  # e.g., 720 - 50 = 670

RULER_TOP_MARGIN = 50       # top of the ruler area
RULER_BOTTOM = CANVAS_HEIGHT  # bottom of the ruler area

# (No camera view is displayed in the main window)

# Column definitions (each column's horizontal bounds)
# Blue: x from 150 to 200 (center ~175)
# Orange: x from 250 to 300 (center ~275)
# Green: x from 350 to 400 (center ~375)
COLUMN_BOUNDS = {
    "Blue":   (150, 200),
    "Orange": (250, 300),
    "Green":  (350, 400)
}

# Mapping ranges for displayed values:
# Blue:    top → 600, bottom → 0  
# Orange:  top → 900, bottom → 600  
# Green:   top → 1200, bottom → 900
BLUE_MIN, BLUE_MAX = 0, 600
ORANGE_MIN, ORANGE_MAX = 600, 900
GREEN_MIN, GREEN_MAX = 900, 1200

# HSV Ranges for Color Detection (adjust as needed)
def load_hsv_ranges():
    if os.path.exists('HSV.data'):
        data = np.load('HSV.data')
        return {
            "Blue": {
                "lower": data['blue_lower'],
                "upper": data['blue_upper'],
                "draw_color": (255, 0, 0)
            },
            "Orange": {
                "lower": data['orange_lower'],
                "upper": data['orange_upper'],
                "draw_color": (0, 165, 255)
            },
            "Green": {
                "lower": data['green_lower'],
                "upper": data['green_upper'],
                "draw_color": (0, 255, 0)
            },
        }
    else:
        return {
            "Blue": {
                "lower": np.array([94, 80, 2]),
                "upper": np.array([126, 255, 255]),
                "draw_color": (255, 0, 0)
            },
            "Orange": {
                "lower": np.array([4, 100, 20]),
                "upper": np.array([25, 255, 255]),
                "draw_color": (0, 165, 255)
            },
            "Green": {
                "lower": np.array([23, 42, 0]),
                "upper": np.array([100, 255, 255]),
                "draw_color": (0, 255, 0)
            },
        }

HSV_RANGES = load_hsv_ranges()

# Morphological operation kernel
KERNEL = np.ones((5, 5), np.uint8)

# Calibration for the cropped image:
# The cropped camera view is 352 pixels tall.
# In our setup, raw ball Y values never go below about 256,
# so we force any raw ball Y below 256 to 256 so that our mapping uses the full column.
DETECTION_Y_MIN = 256
DETECTION_Y_MAX = 352

# =============================================================================
# RFID Reader Window Class
# =============================================================================
def write_to_excel(data, is_header=False):
    gdrive_path = "/home/pi/googledrive/data.xlsx"
    temp_path = "/home/pi/temp_data.xlsx"
    
    try:
        # First, copy existing file from Google Drive to local storage if it exists
        if os.path.exists(gdrive_path):
            import shutil
            shutil.copy2(gdrive_path, temp_path)
            
            # Open the temporary file and append data
            book = load_workbook(temp_path)
            sheet1 = book['Sheet1']
            
            # Create new row as a list
            if len(data) == 1:  # Card ID only
                new_row = [data[0]]
            else:  # Full data row
                new_row = [data[0], data[1], data[2], data[3], data[4]]
            
            # Append to Sheet1
            sheet1.append(new_row)
            
            # Save temporary file
            book.save(temp_path)
        else:
            # Create new workbook with two sheets
            book = Workbook()
            
            # Set up Sheet1 for data
            sheet1 = book.active
            sheet1.title = 'Sheet1'
            
            # Add headers to Sheet1
            if len(data) == 1:
                headers = ['Card ID']
                new_row = [data[0]]
            else:
                headers = ['Card ID', 'Timestamp', 'Blue Value', 'Orange Value', 'Green Value']
                new_row = [data[0], data[1], data[2], data[3], data[4]]
            
            sheet1.append(headers)
            sheet1.append(new_row)
            
            # Create Sheet2 for manual modifications
            sheet2 = book.create_sheet('Sheet2')
            
            # Add column chart to Sheet1
            chart = BarChart()
            chart.type = "col"
            chart.style = 10  # Use a nice style
            chart.title = "Values Over Time"
            chart.x_axis.title = "Timestamp"
            chart.y_axis.title = "Values"
            
            # Create data references for the chart
            # Categories (X-axis) - Timestamp column (B1:B25)
            cats = Reference(sheet1, min_col=2, min_row=1, max_row=25)
            
            # Values (Series) - Blue, Orange, and Green values (C1:E25)
            values = Reference(sheet1, min_col=3, max_col=5, min_row=1, max_row=25)
            
            # Add the data to the chart
            chart.add_data(values, titles_from_data=True)  # Use first row as series names
            chart.set_categories(cats)
            
            # Add the chart to the worksheet
            sheet1.add_chart(chart, "G2")  # Position the chart at cell G2
            
            # Save as temporary file first
            book.save(temp_path)
        
        # Now copy the temporary file back to Google Drive
        import shutil
        if os.path.exists(temp_path):
            shutil.copy2(temp_path, gdrive_path)
            os.remove(temp_path)  # Clean up temp file
            
        return True
        
    except Exception as e:
        # Clean up on error
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except:
                pass
        messagebox.showerror("File Error", f"Could not save data:\n{e}")
        return False

class RFIDReaderWindow:
    def __init__(self, root):
        self.root = root
        self.root.title("RFID Card Reader")
        self.root.configure(bg=BACKGROUND_COLOR)
        self.root.geometry("400x150")
        self.label = tk.Label(root, text="Please scan your RFID card:", font=(FONT_NAME, 16),
                              bg=BACKGROUND_COLOR, fg=TEXT_COLOR)
        self.label.pack(pady=20)
        self.entry = tk.Entry(root, font=(FONT_NAME, 16))
        self.entry.pack(pady=10)
        self.entry.focus_set()
        self.entry.bind("<Return>", self.process_rfid)
        
    def process_rfid(self, event):
        card_number = self.entry.get().strip()
        if card_number:
            self.root.destroy()
            main_root = tk.Tk()
            app = RespiratoryTherapyApp(main_root, card_number)
            main_root.mainloop()

# =============================================================================
# Main Application Class: Real-Time Ball Detection Indicators
# =============================================================================
class RespiratoryTherapyApp:
    def __init__(self, root, card_id):
        self.root = root
        self.card_id = card_id  # RFID card number passed from RFIDReaderWindow
        self.root.title("Respiratory Therapy Device")
        self.root.configure(bg=BACKGROUND_COLOR)
        # Set the main window to full screen
        self.root.attributes("-fullscreen", True)

        # Add running flag for safe shutdown
        self.running = True

        # Dictionary to store detected ball positions (per color)
        self.ball_positions = {key: None for key in HSV_RANGES.keys()}
        self.last_frame_height = DETECTION_Y_MAX

        # Flag to prevent multiple confirmation windows per event.
        self.confirmation_shown = False
        
        # Initialize camera
        self.init_camera()
        
        if not self.picam2:
            messagebox.showerror("Error", "Could not initialize camera. Please restart the application.")
            self.on_closing()
            return

        self.setup_header()
        self.setup_main_panel()

        self.root.after(0, self.update_frame)
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def init_camera(self):
        try:
            self.picam2 = Picamera2()
            preview_config = self.picam2.create_preview_configuration(
                main={"size": (640, 480), "format": "RGB888"}
            )
            self.picam2.configure(preview_config)
            self.picam2.start()
            return True
        except Exception as e:
            print("Error initializing camera:", e)
            self.picam2 = None
            return False

    # -------------------------------
    # GUI Setup for Main Application
    # -------------------------------
    def setup_header(self):
        self.header_frame = tk.Frame(self.root, bg=BACKGROUND_COLOR, height=HEADER_HEIGHT)
        self.header_frame.pack(side="top", fill="x", pady=(0, 5))
        self.header_frame.pack_propagate(False)
        try:
            logo_left = Image.open("qstss.png")
            logo_right = Image.open("moe.png")
        except Exception as e:
            print("Error loading logo images:", e)
            logo_left = Image.new("RGB", (60, 60), color="white")
            logo_right = Image.new("RGB", (60, 60), color="white")
        logo_left = logo_left.resize((80, 50))
        logo_right = logo_right.resize((80, 50))
        self.logo_left_img = ImageTk.PhotoImage(logo_left)
        self.logo_right_img = ImageTk.PhotoImage(logo_right)
        tk.Label(self.header_frame, image=self.logo_left_img, bg=BACKGROUND_COLOR).pack(side="left", padx=5)
        tk.Label(self.header_frame, text="Respiratory Therapy Device", font=(FONT_NAME, 18, "bold"),
                 bg=BACKGROUND_COLOR, fg=TEXT_COLOR).pack(side="left", padx=10)
        tk.Label(self.header_frame, image=self.logo_right_img, bg=BACKGROUND_COLOR).pack(side="right", padx=5)

    def setup_main_panel(self):
        self.main_frame = tk.Frame(self.root, bg=BACKGROUND_COLOR, width=CANVAS_WIDTH, height=CANVAS_HEIGHT)
        self.main_frame.pack(side="top", fill="both", expand=True)
        self.main_frame.pack_propagate(False)
        self.canvas = tk.Canvas(self.main_frame, width=CANVAS_WIDTH, height=CANVAS_HEIGHT,
                                bg=BACKGROUND_COLOR, highlightthickness=0)
        self.canvas.pack(fill="both", expand=True)

        for col in COLUMN_BOUNDS.values():
            x0, x1 = col
            self.canvas.create_rectangle(x0, RULER_TOP_MARGIN, x1, CANVAS_HEIGHT, outline="white", width=2)

        self.draw_column_rulers()

        radius = 20
        self.blue_center_x = (COLUMN_BOUNDS["Blue"][0] + COLUMN_BOUNDS["Blue"][1]) // 2
        self.orange_center_x = (COLUMN_BOUNDS["Orange"][0] + COLUMN_BOUNDS["Orange"][1]) // 2
        self.green_center_x = (COLUMN_BOUNDS["Green"][0] + COLUMN_BOUNDS["Green"][1]) // 2

        # Start with indicators at the bottom and forced displayed values to 0.
        self.blue_default_y = RULER_BOTTOM
        self.orange_default_y = RULER_BOTTOM
        self.green_default_y = RULER_BOTTOM

        self.blue_circle = self.canvas.create_oval(
            self.blue_center_x - radius, self.blue_default_y - radius,
            self.blue_center_x + radius, self.blue_default_y + radius,
            fill="white", outline="blue", width=3)
        self.orange_circle = self.canvas.create_oval(
            self.orange_center_x - radius, self.orange_default_y - radius,
            self.orange_center_x + radius, self.orange_default_y + radius,
            fill="white", outline="orange", width=3)
        self.green_circle = self.canvas.create_oval(
            self.green_center_x - radius, self.green_default_y - radius,
            self.green_center_x + radius, self.green_default_y + radius,
            fill="white", outline="green", width=3)

        self.blue_percent_text = self.canvas.create_text(self.blue_center_x, 30, text="Blue: 0", font=(FONT_NAME, 12), fill="white")
        self.orange_percent_text = self.canvas.create_text(self.orange_center_x, 30, text="Orange: 0", font=(FONT_NAME, 12), fill="white")
        self.green_percent_text = self.canvas.create_text(self.green_center_x, 30, text="Green: 0", font=(FONT_NAME, 12), fill="white")

    def draw_column_rulers(self):
        total_height = CANVAS_HEIGHT - RULER_TOP_MARGIN
        tick_interval = 50
        for col_name, bounds in COLUMN_BOUNDS.items():
            x0, _ = bounds
            if col_name == "Blue":
                col_min, col_max = BLUE_MIN, BLUE_MAX
            elif col_name == "Orange":
                col_min, col_max = ORANGE_MIN, ORANGE_MAX
            elif col_name == "Green":
                col_min, col_max = GREEN_MIN, GREEN_MAX
            else:
                continue
            self.canvas.create_line(x0, RULER_TOP_MARGIN, x0, CANVAS_HEIGHT, width=2, fill="white")
            for value in range(col_min, col_max + 1, tick_interval):
                y = CANVAS_HEIGHT - ((value - col_min) / (col_max - col_min)) * total_height
                self.canvas.create_line(x0, y, x0 + 20, y, width=2, fill="white")
                self.canvas.create_text(x0 - 10, y, text=str(value), font=(FONT_NAME, 10), fill="white")

    @staticmethod
    def process_mask(mask):
        mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, KERNEL, iterations=2)
        mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, KERNEL, iterations=2)
        return mask

    def detect_ball(self, mask, label, draw_color, frame):
        mask = self.process_mask(mask)
        contours, _ = cv2.findContours(mask.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        pos = None
        if contours:
            c = max(contours, key=cv2.contourArea)
            ((x, y), radius) = cv2.minEnclosingCircle(c)
            if radius > 10:
                cv2.circle(frame, (int(x), int(y)), int(radius), draw_color, 2)
                cv2.putText(frame, label, (int(x - radius), int(y - radius)),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, draw_color, 2)
                pos = (int(x), int(y))
                
                # Debug: Show the mask for green ball
                if label == "Green":
                    cv2.imshow("Green Mask", mask)
                    
        self.ball_positions[label] = pos
        return frame

    def get_canvas_y(self, ball_y):
        """
        Map the raw ball y-coordinate (from the cropped image) to a canvas y-coordinate.
        Using the calibration range [DETECTION_Y_MIN, DETECTION_Y_MAX] so that:
          ball_y <= DETECTION_Y_MIN maps to RULER_TOP_MARGIN (top of column)
          ball_y = DETECTION_Y_MAX maps to CANVAS_HEIGHT (bottom of column)
        """
        clamped_y = max(ball_y, DETECTION_Y_MIN)
        normalized = (clamped_y - DETECTION_Y_MIN) / (DETECTION_Y_MAX - DETECTION_Y_MIN)
        normalized = max(0, min(normalized, 1))
        return RULER_TOP_MARGIN + normalized * (CANVAS_HEIGHT - RULER_TOP_MARGIN)

    def update_ball_indicators(self):
        """
        Update ball indicator positions (confined to their columns) and compute mapped values.
        Mapping for each column:
          Blue:    y = RULER_TOP_MARGIN maps to 600, y = CANVAS_HEIGHT maps to 0.
          Orange:  y = RULER_TOP_MARGIN maps to 900, y = CANVAS_HEIGHT maps to 600.
          Green:   y = RULER_TOP_MARGIN maps to 1200, y = CANVAS_HEIGHT maps to 900.
        If no ball is detected in a column, the displayed value is forced to 0.
        """
        def get_new_center(label):
            if self.ball_positions[label]:
                return self.get_canvas_y(self.ball_positions[label][1])
            else:
                return CANVAS_HEIGHT

        radius = 20
        blue_y = get_new_center("Blue")
        orange_y = get_new_center("Orange")
        green_y = get_new_center("Green")

        self.canvas.coords(self.blue_circle,
                           (COLUMN_BOUNDS["Blue"][0] + COLUMN_BOUNDS["Blue"][1]) // 2 - radius,
                           blue_y - radius,
                           (COLUMN_BOUNDS["Blue"][0] + COLUMN_BOUNDS["Blue"][1]) // 2 + radius,
                           blue_y + radius)
        self.canvas.itemconfig(self.blue_circle, fill="blue" if self.ball_positions["Blue"] else "white")

        self.canvas.coords(self.orange_circle,
                           (COLUMN_BOUNDS["Orange"][0] + COLUMN_BOUNDS["Orange"][1]) // 2 - radius,
                           orange_y - radius,
                           (COLUMN_BOUNDS["Orange"][0] + COLUMN_BOUNDS["Orange"][1]) // 2 + radius,
                           orange_y + radius)
        self.canvas.itemconfig(self.orange_circle, fill="orange" if self.ball_positions["Orange"] else "white")

        self.canvas.coords(self.green_circle,
                           (COLUMN_BOUNDS["Green"][0] + COLUMN_BOUNDS["Green"][1]) // 2 - radius,
                           green_y - radius,
                           (COLUMN_BOUNDS["Green"][0] + COLUMN_BOUNDS["Green"][1]) // 2 + radius,
                           green_y + radius)
        self.canvas.itemconfig(self.green_circle, fill="green" if self.ball_positions["Green"] else "white")

        total_height = CANVAS_HEIGHT - RULER_TOP_MARGIN
        blue_raw = (CANVAS_HEIGHT - blue_y) / total_height * (BLUE_MAX - BLUE_MIN) + BLUE_MIN
        orange_raw = (CANVAS_HEIGHT - orange_y) / total_height * (ORANGE_MAX - ORANGE_MIN) + ORANGE_MIN
        green_raw = (CANVAS_HEIGHT - green_y) / total_height * (GREEN_MAX - GREEN_MIN) + GREEN_MIN

        blue_value = int(round(blue_raw)) if self.ball_positions["Blue"] else 0
        orange_value = int(round(orange_raw)) if self.ball_positions["Orange"] else 0
        green_value = int(round(green_raw)) if self.ball_positions["Green"] else 0

        self.canvas.itemconfig(self.blue_percent_text, text=f"Blue: {blue_value:d}")
        self.canvas.itemconfig(self.orange_percent_text, text=f"Orange: {orange_value:d}")
        self.canvas.itemconfig(self.green_percent_text, text=f"Green: {green_value:d}")

        # Check if any indicator has reached its maximum value.
        if (self.ball_positions["Blue"] and blue_value == BLUE_MAX) or \
           (self.ball_positions["Orange"] and orange_value == ORANGE_MAX) or \
           (self.ball_positions["Green"] and green_value == GREEN_MAX):
            if not self.confirmation_shown:
                self.confirmation_shown = True
                self.show_confirmation_window(blue_value, orange_value, green_value)

    def show_confirmation_window(self, blue_value, orange_value, green_value):
        top = tk.Toplevel(self.root)
        top.title("Confirm")
        top.geometry("300x200")
        tk.Label(top, text="Repeat or Finished?", font=(FONT_NAME, 14), bg=BACKGROUND_COLOR, fg=TEXT_COLOR).pack(pady=20)
        btn_frame = tk.Frame(top, bg=BACKGROUND_COLOR)
        btn_frame.pack(pady=10)
        
        # Create progress bar (hidden initially)
        progress_frame = tk.Frame(top, bg=BACKGROUND_COLOR)
        progress_frame.pack(pady=10)
        progress_label = tk.Label(progress_frame, text="Saving data...", font=(FONT_NAME, 10), 
                                bg=BACKGROUND_COLOR, fg=TEXT_COLOR)
        progress_bar = ttk.Progressbar(progress_frame, length=200, mode='determinate')
        
        def show_progress():
            progress_label.pack()
            progress_bar.pack(pady=5)
            # Update progress bar
            for i in range(101):
                progress_bar['value'] = i
                top.update_idletasks()
                time.sleep(0.02)
        
        def repeat():
            top.destroy()
            self.confirmation_shown = False
        
        def finish():
            # Show progress bar
            show_progress()
            
            # Save data
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            data = [self.card_id, now, blue_value, orange_value, green_value]
            if write_to_excel(data):
                # Properly cleanup camera
                if hasattr(self, 'picam2') and self.picam2 is not None:
                    self.picam2.stop()
                    self.picam2.close()
                top.destroy()
                self.running = False
                self.root.destroy()
                new_rfid_root = tk.Tk()
                new_rfid_app = RFIDReaderWindow(new_rfid_root)
                new_rfid_root.mainloop()
        
        tk.Button(btn_frame, text="Repeat", font=(FONT_NAME, 12), command=repeat).pack(side="left", padx=10)
        tk.Button(btn_frame, text="Finished", font=(FONT_NAME, 12), command=finish).pack(side="right", padx=10)

    def update_frame(self):
        if not self.running:
            return
        if self.picam2 is None:
            print("Camera not initialized.")
            self.root.after(10, self.update_frame)
            return

        frame = self.picam2.capture_array()
        if frame is None:
            print("Error: Could not capture frame from camera.")
            self.root.after(10, self.update_frame)
            return

        frame = cv2.flip(frame, 1)  # Mirror effect

        cropped_frame = frame[0:352, 116:430]
        self.last_frame_height = DETECTION_Y_MAX

        blurred = cv2.GaussianBlur(cropped_frame, (11, 11), 0)
        hsv = cv2.cvtColor(blurred, cv2.COLOR_BGR2HSV)
        
        # Debug: Show HSV image
        cv2.imshow("HSV Image", hsv)
        
        processed_frame = np.zeros_like(cropped_frame)
        for label, settings in HSV_RANGES.items():
            mask = cv2.inRange(hsv, settings["lower"], settings["upper"])
            processed_frame = self.detect_ball(mask, label, settings["draw_color"], processed_frame)

        self.update_ball_indicators()
        if self.running:
            self.root.after(10, self.update_frame)

    def on_closing(self):
        self.running = False
        if hasattr(self, 'picam2') and self.picam2 is not None:
            self.picam2.stop()
            self.picam2.close()
        self.root.destroy()


if __name__ == "__main__":
    # Check for running instances first
    if not check_and_create_lock():
        sys.exit(1)
        
    # Launch RFID reader window first.
    try:
        rfid_root = tk.Tk()
        rfid_app = RFIDReaderWindow(rfid_root)
        rfid_root.mainloop()
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")
    finally:
        cleanup_lock()
